import os
from typing import List, Any

from openpyxl import load_workbook


XLSX_FILE = os.getenv("XLSX_FILE", os.path.join("data", "archive", "output.xlsx"))


def first_nonempty_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return i
    return 1


def summarize_sheet(ws) -> None:
    header_row_idx = first_nonempty_row(ws)
    headers: List[str] = [
        (str(c).strip() if c is not None else f"col_{i+1}")
        for i, c in enumerate(next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)))
    ]

    # Count rows after header
    total_rows = 0
    for _ in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        total_rows += 1

    print(f"Sheet: {ws.title}")
    print(f"- Header row index: {header_row_idx}")
    print(f"- Columns ({len(headers)}): {headers}")
    print(f"- Data rows: {total_rows}")

    # Sample first 5 rows
    print("- Sample rows:")
    shown = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        obj = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        print("  ", obj)
        shown += 1
        if shown >= 5:
            break


def main() -> None:
    if not os.path.exists(XLSX_FILE):
        raise SystemExit(f"File not found: {XLSX_FILE}")

    wb = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    print(f"Workbook: {XLSX_FILE}")
    print(f"Sheets: {wb.sheetnames}")

    # Summarize first sheet by default
    ws = wb[wb.sheetnames[0]]
    summarize_sheet(ws)


if __name__ == "__main__":
    main()

