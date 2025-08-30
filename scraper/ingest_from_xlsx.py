import os
import ast
from datetime import datetime, timezone
from typing import Any, Dict, Iterable

from openpyxl import load_workbook
from elasticsearch import Elasticsearch, helpers


ELASTICSEARCH_URL = os.getenv("ELASTICSEARCH_URL", "http://localhost:9200")
ES_INDEX = os.getenv("ES_INDEX", "products")
XLSX_FILE = os.getenv("XLSX_FILE", os.path.join("data", "archive", "output.xlsx"))


def es_client() -> Elasticsearch:
    return Elasticsearch([ELASTICSEARCH_URL])


def ensure_index(es: Elasticsearch) -> None:
    if es.indices.exists(index=ES_INDEX):
        return
    es.indices.create(
        index=ES_INDEX,
        body={
            "settings": {"analysis": {"analyzer": {"default": {"type": "standard"}}}},
            "mappings": {
                "properties": {
                    "title": {"type": "text", "fields": {"raw": {"type": "keyword"}}},
                    "description": {"type": "text"},
                    "category": {"type": "keyword"},
                    "price": {"type": "float"},
                    "url": {"type": "keyword", "index": False},
                    "image_url": {"type": "keyword", "index": False},
                    "source": {"type": "keyword"},
                    "source_id": {"type": "keyword"},
                    "brand": {"type": "keyword"},
                    "average_rating": {"type": "float"},
                    "out_of_stock": {"type": "boolean"},
                    "ingested_at": {"type": "date"}
                }
            },
        },
    )


def to_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).replace(",", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def parse_images(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value if v]
    s = str(value).strip()
    try:
        v = ast.literal_eval(s)
        if isinstance(v, list):
            return [str(x) for x in v if x]
    except Exception:
        pass
    return []


def transform(row: Dict[str, Any]) -> Dict[str, Any]:
    # Some rows in the XLSX have description/discount flipped; detect and swap.
    desc = row.get("description")
    disc = row.get("discount")
    if isinstance(desc, str) and desc.strip().endswith("% off") and isinstance(disc, str) and len(disc) > len(desc):
        desc, disc = disc, desc

    category = row.get("sub_category") or row.get("category")
    price = to_float(row.get("selling_price")) or to_float(row.get("actual_price"))
    images = parse_images(row.get("images"))
    image_url = images[0] if images else None
    pid = str(row.get("pid") or row.get("_id"))
    now_iso = datetime.now(timezone.utc).isoformat()

    return {
        "title": row.get("title"),
        "description": desc,
        "category": category,
        "price": price,
        "url": row.get("url"),
        "image_url": image_url,
        "source": "flipkart",
        "source_id": pid,
        "brand": row.get("brand"),
        "average_rating": to_float(row.get("average_rating")),
        "out_of_stock": bool(row.get("out_of_stock", False)),
        "ingested_at": now_iso,
    }


def iter_rows() -> Iterable[Dict[str, Any]]:
    if not os.path.exists(XLSX_FILE):
        raise SystemExit(f"File not found: {XLSX_FILE}")
    wb = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Read header
    header = [c.value if c.value is not None else f"col_{i+1}" for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))) ]
    # Stream rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        obj = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        yield obj


def actions(rows: Iterable[Dict[str, Any]]):
    for row in rows:
        doc = transform(row)
        if not doc.get("source_id"):
            continue
        yield {
            "_index": ES_INDEX,
            "_id": f"flipkart-{doc['source_id']}",
            "_op_type": "index",
            "_source": doc,
        }


def main() -> None:
    es = es_client()
    ensure_index(es)
    ok, failed = helpers.bulk(es, actions(iter_rows()), stats_only=True, chunk_size=1000)
    print(f"Bulk result: indexed={ok}, errors={failed}")


if __name__ == "__main__":
    main()

